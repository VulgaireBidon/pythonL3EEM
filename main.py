from openpyxl import Workbook
from openpyxl import load_workbook

# charger le classeur 'sample' dans variable wb
wb = load_workbook("sample.xlsx")

# ouvrir la feuille 'Feuil1' dans variable wb
ws = wb["Feuil1"]

# accès à la valeur d'une cellule
a = ws["A2"].value
print(a)

# accès à la valeur d'une cellule
b = ws["B2"].value
print(b)

# ecrire une valeur dans une cellule C2
print(a*b)
ws["C2"].value = a * b

# enregistrer les modificiation dans le classseur 'sampleNew'
wb.save("ex22a-1/sampleAfter.xlsx")

print("Merci. Ouvrez SampleAfter.xlsx")
